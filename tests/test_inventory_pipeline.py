"""
End-to-end test for src.inventory.pipeline

Builds one realistic xlsx file with a mix of valid and invalid rows
(mirroring the actual failure modes called out in the project spec)
and verifies the full read -> validate -> write pipeline produces
correct TXT + quarantine outputs, with rows_read == rows_accepted +
rows_quarantined always holding (nothing is ever lost).
"""

import csv

import openpyxl

from src.inventory.pipeline import process_inventory_file
from src.inventory.schema import EXPECTED_COLUMNS


def _build_realistic_test_file(path):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # Decoy metadata sheets, exactly like real Baapstore files.
    workbook.create_sheet("Instructions").append(["Do not edit."])
    workbook.create_sheet("icons")

    template = workbook.create_sheet("Template")
    template.append(EXPECTED_COLUMNS)
    # 1. Perfectly valid row.
    template.append(["GOOD-SKU-1", None, None, None, 10, None, None, None])
    # 2. Valid row with a float-artifact quantity (10.0) -- must clean to "10".
    template.append(["GOOD-SKU-2", None, None, None, 10.0, None, None, None])
    # 3. Negative quantity -- must be quarantined.
    template.append(["BAD-NEGATIVE", None, None, None, -5, None, None, None])
    # 4. SKU over the 40-byte limit -- must be quarantined.
    template.append(["X" * 45, None, None, None, 3, None, None, None])
    # 5 & 6. Duplicate SKU -- both must be quarantined.
    template.append(["DUPLICATED", None, None, None, 1, None, None, None])
    template.append(["DUPLICATED", None, None, None, 2, None, None, None])
    # 7. Missing SKU -- must be quarantined.
    template.append([None, None, None, None, 7, None, None, None])
    # 8. Missing quantity -- must be quarantined.
    template.append(["NO-QTY-SKU", None, None, None, None, None, None, None])

    workbook.save(path)


def test_full_pipeline_end_to_end(tmp_path):
    xlsx_path = str(tmp_path / "Amazon_Bulk_Daily_Quantity_Update.xlsx")
    _build_realistic_test_file(xlsx_path)

    output_txt_path = str(tmp_path / "output.txt")
    quarantine_csv_path = str(tmp_path / "quarantine_report.csv")

    summary = process_inventory_file(
        xlsx_path=xlsx_path,
        output_txt_path=output_txt_path,
        quarantine_csv_path=quarantine_csv_path,
        timestamp="2026-08-21T12:00:00Z",
    )

    # Nothing is ever lost: every row read is accounted for.
    assert summary.rows_read == 8
    assert summary.rows_accepted == 2
    assert summary.rows_quarantined == 6
    assert summary.rows_accepted + summary.rows_quarantined == summary.rows_read

    # The TXT file contains exactly the 2 good rows, correctly cleaned.
    with open(output_txt_path, "r", encoding="utf-8") as fh:
        lines = fh.read().splitlines()

    assert lines[0] == "\t".join(EXPECTED_COLUMNS)
    assert len(lines) == 3  # header + 2 valid rows
    assert "GOOD-SKU-1\t\t\t\t10\t\t\t" in lines
    assert "GOOD-SKU-2\t\t\t\t10\t\t\t" in lines  # NOT "10.0"

    # The quarantine report accounts for all 6 rejected rows with real reasons.
    with open(quarantine_csv_path, "r", encoding="utf-8", newline="") as fh:
        quarantined = list(csv.DictReader(fh))

    assert len(quarantined) == 6
    reasons = {row["sku"]: row["reason"] for row in quarantined}
    assert reasons["BAD-NEGATIVE"] == "negative_quantity"
    assert reasons["X" * 45] == "sku_exceeds_byte_limit"
    assert reasons["NO-QTY-SKU"] == "missing_quantity"
    duplicate_rows = [row for row in quarantined if row["sku"] == "DUPLICATED"]
    assert len(duplicate_rows) == 2

    # Every quarantine row carries the source file and run timestamp.
    for row in quarantined:
        assert row["source_file"] == "Amazon_Bulk_Daily_Quantity_Update.xlsx"
        assert row["timestamp"] == "2026-08-21T12:00:00Z"


def test_pipeline_defaults_timestamp_to_now_when_not_provided(tmp_path):
    """If no timestamp is injected, the pipeline should still run and produce a real one."""
    xlsx_path = str(tmp_path / "Amazon_Bulk_Daily_Quantity_Update.xlsx")
    _build_realistic_test_file(xlsx_path)

    summary = process_inventory_file(
        xlsx_path=xlsx_path,
        output_txt_path=str(tmp_path / "output.txt"),
        quarantine_csv_path=str(tmp_path / "quarantine.csv"),
    )

    assert summary.rows_read == 8
    