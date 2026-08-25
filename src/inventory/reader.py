"""
Reads the "Template" sheet out of a Baapstore Amazon Inventory Loader
xlsx file.

Deliberately does NOT validate business rules (SKU length, negative
quantities, etc.) -- that's validator.py's job. This module's only
responsibility is: open the file, find the right sheet, confirm its
header row is exactly what we expect, and hand back the raw row data
with original types intact (so validator.py can detect things like
float artifacts that a premature str() conversion here would hide).
"""

import openpyxl

from src.inventory.schema import EXPECTED_COLUMNS, TEMPLATE_SHEET_NAME


def read_template_sheet(xlsx_path: str):
    """
    Read every data row from the "Template" sheet of `xlsx_path`.

    Returns a list of dicts, one per data row, each with:
      - "source_row": the 1-indexed row number in the sheet (as it
        would appear in Excel), for traceability in the quarantine
        report.
      - one key per column in EXPECTED_COLUMNS, holding the raw cell
        value exactly as openpyxl read it (str / int / float / None).

    Raises:
        ValueError: if the file has no "Template" sheet, or if the
            Template sheet's header row doesn't exactly match
            EXPECTED_COLUMNS (same names, same order). We fail loudly
            here rather than guess, per the project's "do not assume
            another worksheet contains the inventory data" rule.
    """
    # read_only=True keeps memory usage low for the ~89,000-row full
    # catalog file. data_only=True returns computed values rather than
    # formula strings (harmless for Baapstore's data-only files, but
    # protects us if a formula ever sneaks in).
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"{xlsx_path!r} has no {TEMPLATE_SHEET_NAME!r} sheet. "
            f"Sheets found: {workbook.sheetnames}"
        )

    sheet = workbook[TEMPLATE_SHEET_NAME]
    row_iterator = sheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iterator)
    except StopIteration:
        raise ValueError(f"{TEMPLATE_SHEET_NAME!r} sheet in {xlsx_path!r} is completely empty.")

    actual_header = list(header_row[: len(EXPECTED_COLUMNS)])
    if actual_header != EXPECTED_COLUMNS:
        raise ValueError(
            f"{TEMPLATE_SHEET_NAME!r} sheet header does not match the expected "
            f"Amazon Inventory Loader columns.\n"
            f"Expected: {EXPECTED_COLUMNS}\n"
            f"Found:    {actual_header}\n"
            "Refusing to guess column meaning -- verify the file format "
            "with Baapstore before proceeding."
        )

    rows = []
    # Row 1 is the header, so data starts at row 2.
    for source_row_number, row_values in enumerate(row_iterator, start=2):
        if row_values is None or all(value is None for value in row_values):
            # Skip fully blank rows (trailing empty rows are common at
            # the end of exported sheets) -- there is nothing to
            # validate or quarantine about a row with no data at all.
            continue

        row_dict = {"source_row": source_row_number}
        for column_name, value in zip(EXPECTED_COLUMNS, row_values):
            row_dict[column_name] = value
        rows.append(row_dict)

    return rows